"""
Reconciliation Script — New Platform vs Legacy Panoply
=======================================================

Purpose:
    Compare aggregated quantity between the new Databricks Lakehouse platform
    and the legacy Panoply Style_selling_dfNEW at (iso_week × vend_id ×
    legacy_channel_group) grain. Generate Leader-friendly Excel report with
    PASS / WARN / FAIL color coding, plus machine-readable JSON summary.

Inputs:
    docs/reconciliation/data/new_platform.csv     (from 01_new_platform_query.sql)
    docs/reconciliation/data/panoply_legacy.csv   (from 02_panoply_legacy_query.sql)

Outputs:
    docs/reconciliation/reports/reconciliation_report.xlsx  (Leader-facing)
    docs/reconciliation/reports/reconciliation_diff.csv     (machine-readable diff)
    docs/reconciliation/reports/reconciliation_summary.json (overall stats)

Usage:
    PowerShell, from the project root:
        cd C:\\Users\\sia.song\\analytics-platform
        python docs/reconciliation/run_reconciliation.py

Dependencies:
    pandas, openpyxl
    Install: pip install pandas openpyxl
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


# =============================================================================
# Configuration
# =============================================================================

# Resolve paths relative to this script (works regardless of where it's run from)
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR / "data"
REPORTS_DIR = SCRIPT_DIR / "reports"

NEW_CSV = DATA_DIR / "new_platform.csv"
PANOPLY_CSV = DATA_DIR / "panoply_legacy.csv"

XLSX_OUT = REPORTS_DIR / "reconciliation_report.xlsx"
DIFF_CSV_OUT = REPORTS_DIR / "reconciliation_diff.csv"
SUMMARY_JSON_OUT = REPORTS_DIR / "reconciliation_summary.json"

# Reconciliation thresholds (independent of notebook 04's DQ thresholds).
# Calibrated for GA4-vs-TW channel taxonomy translation + DST timezone
# correction + TW data freshness — wider than ETL-level DQ thresholds.
PASS_THRESHOLD = 0.02   # |diff%| < 2%  → PASS  (green)
WARN_THRESHOLD = 0.05   # |diff%| < 5%  → WARN  (yellow)
                         # >= 5%         → FAIL  (red)
# Single-side missing (one side has data, other side is empty) is flagged
# separately as MISSING when the existing side has > 10 units (below this,
# it's likely just sparse data noise, not a real discrepancy).
MISSING_UNITS_FLOOR = 10

# Excel cell colors (openpyxl PatternFill hex without leading '#')
COLOR_PASS = "C6EFCE"      # light green
COLOR_WARN = "FFEB9C"      # light yellow
COLOR_FAIL = "FFC7CE"      # light red
COLOR_MISSING = "FFD9B3"   # light orange
COLOR_HEADER = "BDD7EE"    # light blue

JOIN_KEYS = ["iso_year", "iso_week", "vend_id", "legacy_channel_group"]


# =============================================================================
# Core logic
# =============================================================================

def load_csv(path: Path, qty_col: str) -> pd.DataFrame:
    """Load a reconciliation CSV with validation. Returns DataFrame indexed
    on JOIN_KEYS with a single quantity column named `qty_col`."""
    if not path.exists():
        sys.exit(
            f"ERROR: Input file not found: {path}\n"
            f"Run the SQL queries first and save outputs to {DATA_DIR}/"
        )

    df = pd.read_csv(path)

    # Validate required columns
    required = JOIN_KEYS + [qty_col]
    missing = [c for c in required if c not in df.columns]
    if missing:
        sys.exit(
            f"ERROR: {path.name} missing required columns: {missing}\n"
            f"Got columns: {list(df.columns)}"
        )

    # Normalize: strip whitespace from string keys, cast numeric keys to int
    df["iso_year"] = df["iso_year"].astype(int)
    df["iso_week"] = df["iso_week"].astype(int)
    df["vend_id"] = df["vend_id"].astype(str).str.strip()
    df["legacy_channel_group"] = df["legacy_channel_group"].astype(str).str.strip()
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0).astype(int)

    return df[required]


def categorize(row: pd.Series) -> str:
    """Classify a reconciliation row into PASS / WARN / FAIL / MISSING_NEW /
    MISSING_PANOPLY."""
    q_new = row["qty_new"]
    q_pan = row["qty_panoply"]

    # Both empty (shouldn't happen after outer join, but guard)
    if pd.isna(q_new) and pd.isna(q_pan):
        return "PASS"

    # Treat NaN as 0 for diff math, but flag missing-side cases
    q_new_eff = 0 if pd.isna(q_new) else q_new
    q_pan_eff = 0 if pd.isna(q_pan) else q_pan

    # Missing-side flagging (one side empty, other has meaningful volume)
    if q_pan_eff == 0 and q_new_eff > MISSING_UNITS_FLOOR:
        return "MISSING_PANOPLY"
    if q_new_eff == 0 and q_pan_eff > MISSING_UNITS_FLOOR:
        return "MISSING_NEW"

    # Both sides have data (or noise) — compute pct diff
    denom = max(q_pan_eff, q_new_eff, 1)
    pct = abs(q_new_eff - q_pan_eff) / denom

    if pct < PASS_THRESHOLD:
        return "PASS"
    if pct < WARN_THRESHOLD:
        return "WARN"
    return "FAIL"


def reconcile(new_df: pd.DataFrame, pan_df: pd.DataFrame) -> pd.DataFrame:
    """Full outer join + diff calculation + status categorization."""
    merged = new_df.merge(pan_df, on=JOIN_KEYS, how="outer")
    merged["qty_new"] = merged["qty_new"].fillna(0).astype(int)
    merged["qty_panoply"] = merged["qty_panoply"].fillna(0).astype(int)
    merged["diff"] = merged["qty_new"] - merged["qty_panoply"]
    denom = merged[["qty_new", "qty_panoply"]].max(axis=1).clip(lower=1)
    merged["pct_diff"] = (merged["diff"].abs() / denom).round(4)
    merged["status"] = merged.apply(categorize, axis=1)
    return merged.sort_values(JOIN_KEYS).reset_index(drop=True)


def compute_summary(diff_df: pd.DataFrame) -> dict:
    """Aggregate per-status counts and total quantity stats."""
    total_rows = len(diff_df)
    status_counts = diff_df["status"].value_counts().to_dict()
    status_pct = {k: round(v / total_rows, 4) for k, v in status_counts.items()}

    qty_new_total = int(diff_df["qty_new"].sum())
    qty_pan_total = int(diff_df["qty_panoply"].sum())
    overall_diff = qty_new_total - qty_pan_total
    overall_pct = (
        round(abs(overall_diff) / max(qty_pan_total, 1), 4)
        if qty_pan_total > 0 else None
    )

    # Top 10 worst FAIL / MISSING rows for quick attention
    bad = diff_df[diff_df["status"].isin(["FAIL", "MISSING_NEW", "MISSING_PANOPLY"])]
    top_worst = (
        bad.sort_values("pct_diff", ascending=False)
           .head(10)
           .to_dict(orient="records")
    )

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "thresholds": {
            "pass_lt_pct": PASS_THRESHOLD,
            "warn_lt_pct": WARN_THRESHOLD,
            "missing_units_floor": MISSING_UNITS_FLOOR,
        },
        "total_rows": total_rows,
        "status_counts": status_counts,
        "status_pct": status_pct,
        "totals": {
            "qty_new_total": qty_new_total,
            "qty_panoply_total": qty_pan_total,
            "overall_abs_diff": int(abs(overall_diff)),
            "overall_pct_diff": overall_pct,
        },
        "top_10_worst": top_worst,
    }


# =============================================================================
# Excel output
# =============================================================================

def write_excel(diff_df: pd.DataFrame, summary: dict, out_path: Path) -> None:
    """Write a multi-sheet, color-coded XLSX for Leader review."""
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ---------- Sheet 1: Summary ----------
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _write_summary_sheet(ws_sum, summary)

    # ---------- Sheet 2: Full diff ----------
    ws_diff = wb.create_sheet("All_Rows")
    _write_diff_sheet(ws_diff, diff_df, color_status_col=True)

    # ---------- Sheet 3: Failures only ----------
    bad_df = diff_df[diff_df["status"].isin(["FAIL", "MISSING_NEW", "MISSING_PANOPLY"])]
    ws_bad = wb.create_sheet("Needs_Attention")
    _write_diff_sheet(ws_bad, bad_df, color_status_col=True)

    wb.save(out_path)


def _write_summary_sheet(ws, summary: dict) -> None:
    """Render the summary sheet — high-level stats Leader sees first."""
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    center = Alignment(horizontal="center", vertical="center")

    ws["A1"] = "Reconciliation Report — Slice 1"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Generated: {summary['generated_at']}"
    ws["A3"] = (
        f"Grain: iso_year × iso_week × vend_id × legacy_channel_group  |  "
        f"Window: last 7 complete days"
    )

    # Status breakdown table
    ws["A5"] = "Status"
    ws["B5"] = "Rows"
    ws["C5"] = "% of total"
    for cell in ("A5", "B5", "C5"):
        ws[cell].font = bold
        ws[cell].fill = header_fill
        ws[cell].alignment = center

    row = 6
    color_map = {
        "PASS": COLOR_PASS, "WARN": COLOR_WARN, "FAIL": COLOR_FAIL,
        "MISSING_NEW": COLOR_MISSING, "MISSING_PANOPLY": COLOR_MISSING,
    }
    for status in ["PASS", "WARN", "FAIL", "MISSING_NEW", "MISSING_PANOPLY"]:
        count = summary["status_counts"].get(status, 0)
        pct = summary["status_pct"].get(status, 0)
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=f"{pct:.2%}")
        fill = PatternFill("solid", fgColor=color_map[status])
        for col in range(1, 4):
            ws.cell(row=row, column=col).fill = fill
        row += 1

    # Total quantity comparison
    row += 1
    ws.cell(row=row, column=1, value="Metric").font = bold
    ws.cell(row=row, column=2, value="Value").font = bold
    for col in (1, 2):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).alignment = center
    row += 1
    t = summary["totals"]
    rows = [
        ("Total qty (new platform)", t["qty_new_total"]),
        ("Total qty (Panoply legacy)", t["qty_panoply_total"]),
        ("Overall |diff|", t["overall_abs_diff"]),
        ("Overall % diff", f"{t['overall_pct_diff']:.2%}" if t["overall_pct_diff"] is not None else "N/A"),
    ]
    for label, val in rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        row += 1

    # Thresholds reference
    row += 2
    ws.cell(row=row, column=1, value="Threshold reference (this report):").font = bold
    row += 1
    th = summary["thresholds"]
    ws.cell(row=row, column=1, value=f"PASS  : |diff%| < {th['pass_lt_pct']:.0%}")
    row += 1
    ws.cell(row=row, column=1, value=f"WARN  : {th['pass_lt_pct']:.0%} ≤ |diff%| < {th['warn_lt_pct']:.0%}")
    row += 1
    ws.cell(row=row, column=1, value=f"FAIL  : |diff%| ≥ {th['warn_lt_pct']:.0%}")
    row += 1
    ws.cell(row=row, column=1, value=f"MISSING: one side ≥ {th['missing_units_floor']} units, other side 0")

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14


def _write_diff_sheet(ws, df: pd.DataFrame, color_status_col: bool = True) -> None:
    """Render a diff DataFrame to a sheet with row-level color coding."""
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    color_map = {
        "PASS": COLOR_PASS, "WARN": COLOR_WARN, "FAIL": COLOR_FAIL,
        "MISSING_NEW": COLOR_MISSING, "MISSING_PANOPLY": COLOR_MISSING,
    }

    if df.empty:
        ws["A1"] = "No rows in this category."
        ws["A1"].font = bold
        return

    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold
        cell.fill = header_fill

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

        # Color the row based on status
        status = getattr(row, "status", None)
        if status and color_status_col:
            fill = PatternFill("solid", fgColor=color_map.get(status, "FFFFFF"))
            for col_idx in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            *(len(str(v)) for v in df[col_name].astype(str).head(50)),
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 30)


# =============================================================================
# Main
# =============================================================================

def main() -> int:
    print(f"[INFO] Reconciliation run starting at {datetime.now().isoformat(timespec='seconds')}")
    print(f"[INFO] Script dir: {SCRIPT_DIR}")

    # Ensure output dir exists
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # Load inputs
    print(f"[INFO] Loading new platform CSV: {NEW_CSV}")
    new_df = load_csv(NEW_CSV, "qty_new")
    print(f"[INFO]   {len(new_df):,} rows")

    print(f"[INFO] Loading Panoply legacy CSV: {PANOPLY_CSV}")
    pan_df = load_csv(PANOPLY_CSV, "qty_panoply")
    print(f"[INFO]   {len(pan_df):,} rows")

    # Reconcile
    print("[INFO] Computing diff...")
    diff_df = reconcile(new_df, pan_df)
    print(f"[INFO]   {len(diff_df):,} reconciliation buckets")

    # Summary
    summary = compute_summary(diff_df)

    # Write outputs
    print(f"[INFO] Writing Excel report: {XLSX_OUT}")
    write_excel(diff_df, summary, XLSX_OUT)

    print(f"[INFO] Writing diff CSV: {DIFF_CSV_OUT}")
    diff_df.to_csv(DIFF_CSV_OUT, index=False)

    print(f"[INFO] Writing summary JSON: {SUMMARY_JSON_OUT}")
    with SUMMARY_JSON_OUT.open("w") as f:
        json.dump(summary, f, indent=2, default=str)

    # Console summary
    print()
    print("=" * 70)
    print("RECONCILIATION SUMMARY")
    print("=" * 70)
    print(f"Total reconciliation buckets : {summary['total_rows']:,}")
    print()
    print("Status breakdown:")
    for status in ["PASS", "WARN", "FAIL", "MISSING_NEW", "MISSING_PANOPLY"]:
        count = summary["status_counts"].get(status, 0)
        pct = summary["status_pct"].get(status, 0)
        print(f"  {status:18s}: {count:>6,}  ({pct:6.2%})")
    print()
    print(f"Total qty (new platform)     : {summary['totals']['qty_new_total']:,}")
    print(f"Total qty (Panoply legacy)   : {summary['totals']['qty_panoply_total']:,}")
    overall_pct = summary["totals"]["overall_pct_diff"]
    if overall_pct is not None:
        print(f"Overall % diff               : {overall_pct:.2%}")
    print("=" * 70)
    print(f"[OK] Reports written to {REPORTS_DIR}/")

    return 0


if __name__ == "__main__":
    sys.exit(main())
